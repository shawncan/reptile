#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
import re
import logging
from bs4 import BeautifulSoup
import openpyxl
import os
import datetime


def getHTMLText(url):
    """
    下载目标网页源码
    """
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36',
    }

    try:

        r = requests.get(url, headers=headers)
        r.raise_for_status()
        r.encoding = 'utf-8'
        return r.text
    except Exception:
        logger.exception("Download HTML Text failed")


def writeExcel(poxy_list):
    """
    把提取出来的信息写入excel
    """
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/agentPool.xlsx'
    title = ['ip', '类型', '验证时间',]

    if not os.path.exists(pata):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"].value = title[0]
        sheet["B1"].value = title[1]
        sheet["C1"].value = title[2]
        workbook.save(pata)

    page_workbook = openpyxl.load_workbook(pata)
    page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0])
    row = page_sheet.max_row
    for i in range(len(poxy_list)):
        page_sheet["A%d" % (row + i + 1)].value = poxy_list[i]['ip']
        page_sheet["B%d" % (row + i + 1)].value = poxy_list[i]['类型']
        page_sheet["C%d" % (row + i + 1)].value = poxy_list[i]['验证时间']
    page_workbook.save(pata)


def extractIp(limit_list, url):
    """
    提取此区间内的代理信息并保存到Excel
    """
    status = True
    proxy_list = []
    html = getHTMLText(url)
    soup = BeautifulSoup(html, 'html.parser')
    tbody = soup.find('table')
    tr = tbody.find_all('tr')
    crawl_capped = limit_list[1]
    crawl_lower_limit = limit_list[0]

    today = datetime.date.today()

    if int(crawl_capped) == 23:
        yesterday = today - datetime.timedelta(days=1)
        current_date = yesterday.strftime("%d")
    else:
        current_date = today.strftime("%d")

    print(current_date)
    for info_fast in tr[1:]:
        proxy_data = {'ip': '', '类型': '', '存活时间': '', '验证时间': '', }

        extrac_info = re.findall(r'<td>.*</td>', str(info_fast))

        # 验证时间
        verification_time = extrac_info[4][4:-5]
        # 验证的小时
        hour = verification_time.split(" ")[1].split(":")[0]
        # 验证的日期
        day = verification_time.split(" ")[0].split("-")[2]

        if int(day) != int(current_date):
            # status = False
            if int(crawl_capped) != 23:
                status = False
            continue

        if int(hour) >= int(crawl_capped):
            if int(day) != int(current_date):
                print("！")
            continue

        if int(crawl_lower_limit) > int(hour):
            status = False
            continue

        judgment_info = re.findall(r'<div class="bar" .*>', str(info_fast))
        # 速度
        speed = judgment_info[0].split('"')[3][:-1]
        # 连接时间
        connect_time = judgment_info[1].split('"')[3][:-1]

        if float(speed) > 3:
            continue
        if float(connect_time) > 1:
            continue

        # 存活时间
        survival_time = extrac_info[3][4:-5]
        # 时间属性
        time_attribute = re.findall(r'([\u4e00-\u9fa5]+)', survival_time)

        if time_attribute[0] != "天":
            continue

        # ip地址
        ip = extrac_info[0][4:-5]
        # 端口
        port = extrac_info[1][4:-5]
        # 代理ip
        poxyIp = ("{ip}:{port}").format(ip=ip, port=port)
        # 类型
        type = extrac_info[2][4:-5]

        proxy_data['ip'] = poxyIp
        proxy_data['类型'] = type
        proxy_data['验证时间'] = verification_time

        proxy_list.append(proxy_data)

    writeExcel(proxy_list)
    print("\r{url}Crawling OK...".format(url=url, end=""))
    return status


if __name__ == '__main__':
    logging.basicConfig(level=logging.WARNING,
                        format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='/Users/wangjiacan/Desktop/代码/log/xicidaili_proxy.txt',
                        filemode='a')

    logger = logging.getLogger()
    crawl_time = [["0", "10"], ["10", "17"], ["17", "23"]]
    crawl_time_num = 0
    enable = True

    now_time = datetime.datetime.now()
    current_time = now_time.strftime("%H")
    if 0 < int(current_time) < 10:
        crawl_time_num = 2
    elif 10 < int(current_time) < 17:
        crawl_time_num = 0
    elif 17 < int(current_time) <= 23:
        crawl_time_num = 1

    label_num = 0
    while enable:
        label_num += 1
        xicidaili_url = "http://www.xicidaili.com/nn/" + str(label_num)
        enable = extractIp(crawl_time[crawl_time_num], xicidaili_url)


