#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import openpyxl
import os


aims_pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/duoban_book_pro.xlsx'
file_pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/duoban_book_pro_wjc.xlsx'

goheavt_list = []
file_conent = []
file_length = 1

print("去重工作开始...")
if not os.path.exists(file_pata):
    workbook = openpyxl.Workbook()
    workbook.save(file_pata)


aims_workbook = openpyxl.load_workbook(aims_pata)
aims_sheet = aims_workbook.get_sheet_by_name(aims_workbook.get_sheet_names()[0])
aims_row = aims_sheet.max_row


for i in range(aims_row):
    goheavt_aims = list(aims_sheet.rows)[i][1].value

    if str(goheavt_aims) in goheavt_list:
        continue
    goheavt_list.append(goheavt_aims)

    for cell in list(aims_sheet.rows)[i]:
        file_conent.append(cell.value)

    file_workbook = openpyxl.load_workbook(file_pata)
    file_sheet = file_workbook.get_sheet_by_name(file_workbook.get_sheet_names()[0])

    file_sheet["A%d" % file_length].value = file_conent[0]
    file_sheet["B%d" % file_length].value = file_conent[1]
    file_sheet["C%d" % file_length].value = file_conent[2]
    file_sheet["D%d" % file_length].value = file_conent[3]
    file_sheet["E%d" % file_length].value = file_conent[4]
    file_sheet["F%d" % file_length].value = file_conent[5]
    file_workbook.save(file_pata)

    file_length += 1
    file_conent.clear()

aims_workbook.save(aims_pata)
print("去重工作结束...")
