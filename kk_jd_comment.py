#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import re
import logging
import openpyxl
import threading
import queue
import datetime
import time
import os
import json
import demjson


def getHTMLText(url, ip):
    """
    下载目标网页源码
    """
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:57.0) Gecko/20100101 Firefox/57.0',
    }
    proxies = {
        "http": "http://" + ip,
        "https": "http://" + ip,
    }

    try:
        r = requests.get(url, headers=headers, proxies=proxies, timeout=20)
        r.raise_for_status()
        r.encoding = 'GBK'
        return r.text
    except Exception:
        logger.exception("Download HTML Text failed")


def commentProcessing(contentInfo):
    """
    处理评论输出中文评论
    """
    symbol_list = re.findall(r"&[a-z]*;", contentInfo)
    if symbol_list:
        for symbol in set(symbol_list):
            if symbol == "&hellip;":
                new_symbol = "..."
            elif symbol == "&ldquo;":
                new_symbol = "'"
            elif symbol == "&rdquo;":
                new_symbol = "'"
            else:
                new_symbol = " "
            new_content = re.sub(symbol, new_symbol, contentInfo)
            contentInfo = new_content
    return contentInfo


def readExcel(proxyInfo_list):
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/agentPool.xlsx'
    page_workbook = openpyxl.load_workbook(pata)
    page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0])
    for row in page_sheet.rows:
        proxyData = {'ip': '', '类型': '', '验证时间': '', }
        ip = row[0].value
        httpType = row[1].value
        checkDtime = row[2].value

        proxyData["ip"] = ip
        proxyData["类型"] = httpType
        proxyData["验证时间"] = checkDtime

        proxyInfo_list.append(proxyData)
    page_workbook.save(pata)


def writrProxyExcel(proxyInfo_list):
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/agentPool.xlsx'
    page_workbook = openpyxl.load_workbook(pata)

    # 删除表
    page_workbook.remove_sheet(page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0]))

    # 新建表
    new_sheet = page_workbook.create_sheet()

    for i in range(len(proxyInfo_list)):
        new_sheet["A%d" % (i + 1)].value = proxyInfo_list[i]['ip']
        new_sheet["B%d" % (i + 1)].value = proxyInfo_list[i]['类型']
        new_sheet["C%d" % (i + 1)].value = proxyInfo_list[i]['验证时间']

    page_workbook.save(pata)


def writeExcel(comment_list):
    """
    把提取出来的信息写入excel
    """
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/jdkk_comment.xlsx'
    title = ['商品类型', '用户名', '会员等级', '评价时间', '评价内容', '追评时间', '追评内容','评价客户端']

    if not os.path.exists(pata):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"].value = title[0]
        sheet["B1"].value = title[1]
        sheet["C1"].value = title[2]
        sheet["D1"].value = title[3]
        sheet["E1"].value = title[4]
        sheet["F1"].value = title[5]
        sheet["G1"].value = title[6]
        sheet["H1"].value = title[7]
        workbook.save(pata)

    page_workbook = openpyxl.load_workbook(pata)
    page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0])
    row = page_sheet.max_row
    for i in range(len(comment_list)):
        page_sheet["A%d" % (row + i + 1)].value = comment_list[i]['商品类型']
        page_sheet["B%d" % (row + i + 1)].value = comment_list[i]['用户名']
        page_sheet["C%d" % (row + i + 1)].value = comment_list[i]['会员等级']
        page_sheet["D%d" % (row + i + 1)].value = comment_list[i]['评价时间']
        page_sheet["E%d" % (row + i + 1)].value = comment_list[i]['评价内容']
        page_sheet["F%d" % (row + i + 1)].value = comment_list[i]['追评时间']
        page_sheet["G%d" % (row + i + 1)].value = comment_list[i]['追评内容']
        page_sheet["H%d" % (row + i + 1)].value = comment_list[i]['评价客户端']
    page_workbook.save(pata)


def getContentExtraction(buyersomments):
    """
    提取评论链接中的信息
    """
    jdkkComment_list = []

    for num in range(len(buyersomments)):
        kkComment_data = {'商品类型': '', '用户名': '', '会员等级': '', '评价时间': '', '评价内容': '','追评时间': '',
                          '追评内容': '', '评价客户端': '',}

        # 商品类型
        productColor = buyersomments[num]["productColor"]
        # 用户名
        nickname = buyersomments[num]["nickname"]
        # 会员等级
        userLevelName = buyersomments[num]["userLevelName"]

        # 第一次评价时间
        creationTime = buyersomments[num]["creationTime"]
        # 第一次评价内容
        content = buyersomments[num]["content"]
        modify_content = commentProcessing(content)

        # 第二次评价信息
        created = ''
        modify_mostcontent = ''
        if "afterUserComment" in buyersomments[num].keys():
            afterUserComment = buyersomments[num]["afterUserComment"]
            # 第二次评价时间
            created = afterUserComment["created"]
            # 第二次评价内容
            hAfterUserComment = afterUserComment["hAfterUserComment"]
            mostcontent = hAfterUserComment["content"]
            modify_mostcontent = commentProcessing(mostcontent)

        # 评价客户端
        userClientShow = buyersomments[num]['userClientShow']

        kkComment_data['商品类型'] = productColor
        kkComment_data['用户名'] = nickname
        kkComment_data['会员等级'] = userLevelName
        kkComment_data['评价时间'] = creationTime
        kkComment_data['评价内容'] = modify_content
        kkComment_data['追评时间'] = created
        kkComment_data['追评内容'] = modify_mostcontent
        kkComment_data['评价客户端'] = userClientShow

        jdkkComment_list.append(kkComment_data)

    writeExcel(jdkkComment_list)


def getEvaluationUrl():
    productId = 4183290
    # 2为中评，1为差评
    score = 1
    page = 0
    enable = True
    proxyInfo = []
    deleteList = 0
    html = ''

    readExcel(proxyInfo)

    while enable:
        evaluationUrl = 'https://sclub.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98vv445&' \
              'productId={productId}&score={score}&sortType=5&page={page}&pageSize=10&isShadowSku=0&rid=0&fold=1'.format\
            (productId=productId, score=score, page=page)

        for num in range(len(proxyInfo)):
            if num == 0:
                continue

            proxyIP = proxyInfo[num]['ip']
            html = getHTMLText(evaluationUrl, proxyIP)

            if html == None:
                deleteList += 1
                proxyInfo.remove(proxyInfo[num])
                continue
            else:
                break

        htmlInfo = html[25:-2]
        returnJosn = json.loads(htmlInfo)
        buyersCommentsInfo = returnJosn['comments']
        if buyersCommentsInfo:
            getContentExtraction(buyersCommentsInfo)
            print("第{page}页评论爬取完成".format(page=page))
        else:
            enable = False
        page += 1

    print(deleteList)
    print(len(proxyInfo))

    writrProxyExcel(proxyInfo)
    print("评论完全爬取完成")


if __name__ == '__main__':
    logging.basicConfig(level=logging.WARNING,
                        format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='/Users/wangjiacan/Desktop/代码/log/kk_jd_comment.txt',
                        filemode='a')

    logger = logging.getLogger()

    getEvaluationUrl()