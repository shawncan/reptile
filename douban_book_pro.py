#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import re
import logging
import openpyxl
import threading
import queue
import datetime
import time
import os


def deleteProxy():
    """
    删除失效代理
    """
    path = '/Users/wangjiacan/Desktop/shawn/爬取资料/ip/ip.txt'

    file_read = open(path, 'r')
    file_content = ''.join(file_read.readlines()[1:])
    file_write = open(path, 'w')
    file_write.write(file_content)
    file_read.close()
    file_write.close()


def getHTMLText(url):
    """
    下载目标网页源码
    """
    path = '/Users/wangjiacan/Desktop/shawn/爬取资料/ip/proxy_ip.txt'
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36',
    }

    error_num = 0
    mutex.acquire()
    file = open(path, 'r')
    file_num = len(file.readlines())
    file.close()
    html = ''
    for i in range(file_num):
        ip_file = open(path, 'r')
        proxy_ip = ip_file.readlines()[0][:-1]
        ip_file.close()

        proxies = {
            "http": "http://" + proxy_ip,
            "https": "http://" + proxy_ip
        }

        try:
            r = requests.get(url, headers=headers, proxies=proxies, timeout=20)
            status = r.status_code
            print('代理：{ip}  网址：{url}  状态：{url_status}'.format(ip=proxy_ip, url=url, url_status=status))
            if int(status) == 403:
                deleteProxy()
                continue
            else:
                r.raise_for_status()
                r.encoding = 'utf-8'
                html = r.text
                soup = BeautifulSoup(html, 'html.parser')
                script = soup.find_all("script", attrs={'type': 'text/javascript'})
                if not script:
                    deleteProxy()
                    continue
                break
        except Exception:
            error_num += 1
            print('网页链接请求失败：第{error_num}次'.format(error_num=error_num))
            if error_num == 3:
                deleteProxy()
                error_num = 0
            continue
    mutex.release()
    return html


def getTagUrl(url):
    """
    提取所有需要爬取信息的网页链接
    """
    tag_list = []
    html = getHTMLText(url)
    soup = BeautifulSoup(html, 'html.parser')
    try:
        article = soup.find(attrs={'class': 'article'})
        div = article.find_all('div')
        tag_list.append(div[2])
        tag_list.append(div[3])
        tag_list.append(div[4])

        for tagCol in tag_list:
            piece = re.findall(r'"/tag/.*"', str(tagCol))
            for tag_url in piece:
                url_queue.put(tag_url[1:-1])
    except Exception:
        logger.exception("Site link extraction failed")


def getHTMLUrl(url, url_list):
    """
    提取所有需要爬取信息的网页链接
    """
    number = 0
    tag_url = 'https://book.douban.com' + url
    url_list.append(url)

    html = getHTMLText(tag_url)
    soup = BeautifulSoup(html, 'html.parser')
    try:
        paginator = soup.find(attrs={'class': 'paginator'})
        page = paginator.find_all('a')[1:-1]
        thispage = paginator.find_all(attrs={'class': 'thispage'})[0].text
        for url in range(len(page)):
            page_num = page[url].text
            if int(page_num) < int(thispage):
                continue
            if number == 4:
                break
            number += 1
            page_url = page[url].attrs['href']
            url_list.append(page_url)
    except Exception:
        logger.exception("Site link extraction failed")


def getContentExtraction(url):
    """
    提取目标网页中的书名、作者、评分、评价人数、评语的信息并保存到Excel
    """
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/duoban_book_pro.xlsx'
    page_url = 'https://book.douban.com' + url
    book_list = []
    status = True
    html = getHTMLText(page_url)
    soup = BeautifulSoup(html, 'html.parser')

    try:
        subject_list = soup.find(attrs={'class': 'subject-list'})
        content = soup.find_all(attrs={'id': 'content'})[0].h1.text.split(" ")[1]
        h2 = subject_list.find_all('h2')
        pub = subject_list.find_all(attrs={'class': 'pub'})
        star_clearfix = subject_list.find_all(attrs={'class': 'star clearfix'})
        pl = subject_list.find_all(attrs={'class': 'pl'})
        
        rel = soup.find_all(attrs={'rel': 'next'})
        pl2 = soup.find_all(attrs={'class': 'pl2'})[0].text

        if not rel:
            status = False

        if pl2 == '没有找到符合条件的图书':
            status = False

        for i in range(len(h2)):
            book_data = {'标签': '', '书名': '', '作者': '', '评分': '', '评价人数': '', '书本链接': '', }

            comment = re.findall(r'\d*人评价', str(pl[i]))[0][:-3]
            if int(comment) < 20000:
                continue

            rating_nums = re.findall(r'\d.\d<', str(star_clearfix[i]))
            if rating_nums:
                score = rating_nums[0][:-1]
            else:
                continue
            if float(score) < 8.5:
                continue

            book_name = h2[i].a.attrs['title']
            author = pub[i].text.split("  ")[6].split("/")[0]
            number_of_comments = re.findall(r'\d*人评价', str(pl[i]))[0]
            book_url = h2[i].a.attrs['href']

            book_data['标签'] = content
            book_data['书名'] = book_name
            book_data['作者'] = author
            book_data['评分'] = score
            book_data['评价人数'] = number_of_comments
            book_data['书本链接'] = book_url

            book_list.append(book_data)

        mutex.acquire()
        page_workbook = openpyxl.load_workbook(pata)
        page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0])
        row = page_sheet.max_row
        for i in range(len(book_list)):
            page_sheet["A%d" % (row + i + 1)].value = book_list[i]['标签']
            page_sheet["B%d" % (row + i + 1)].value = book_list[i]['书名']
            page_sheet["C%d" % (row + i + 1)].value = book_list[i]['作者']
            page_sheet["D%d" % (row + i + 1)].value = book_list[i]['评分']
            page_sheet["E%d" % (row + i + 1)].value = book_list[i]['评价人数']
            page_sheet["F%d" % (row + i + 1)].value = book_list[i]['书本链接']
        page_workbook.save(pata)
        print("\r{url}数据爬取完成...".format(url=page_url, end=""))
        mutex.release()
        return status
    except Exception:
        logger.exception("Failed to extract {link} information".format(link=page_url))


def Start(url):
    page_url_list = []
    enable = True
    wait_url = url
    tag_name = url[5:]
    page_num = 0

    try:
        while enable:
            getHTMLUrl(wait_url, page_url_list)
            wait_url = page_url_list[-1]
            if len(page_url_list) > 5:
                page_url_list.pop()
            for p_url in page_url_list:
                if not enable:
                    break
                enable = getContentExtraction(p_url)
                time.sleep(2)
                page_num += 1
            page_url_list.clear()
        print("豆瓣图书标签:{tag}爬取完成，共爬取{num}个页面...".format(tag=tag_name, num=page_num))
    except Exception:
        logger.exception("Running encountered an error")


if __name__ == '__main__':
    logging.basicConfig(level=logging.WARNING,
                        format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='/Users/wangjiacan/Desktop/代码/log/DBTS_pro.txt',
                        filemode='a')

    file_pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/duoban_book_pro.xlsx'
    title = ['标签', '书名', '作者', '评分', '评价人数', '书本链接']
    logger = logging.getLogger()
    url_queue = queue.Queue()
    threads = []
    threads_url = []

    print("豆瓣爬虫爬取开始...")
    start = datetime.datetime.now()
    if not os.path.exists(file_pata):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"].value = title[0]
        sheet["B1"].value = title[1]
        sheet["C1"].value = title[2]
        sheet["D1"].value = title[3]
        sheet["E1"].value = title[4]
        sheet["F1"].value = title[5]
        workbook.save(file_pata)

    mutex = threading.Lock()
    initial_url = 'https://book.douban.com/tag/?view=type'
    getTagUrl(initial_url)

    while url_queue.qsize() != 0:
        if url_queue.qsize() < 6:
            num = url_queue.qsize()
        else:
            num = 6
        for i in range(num):
            threads_url.append(url_queue.get())

        for t in range(len(threads_url)):
            print("豆瓣图书标签:{tag}开始爬取...".format(tag=threads_url[t][5:]))
            threads.append(threading.Thread(target=Start, args=(threads_url[t], )))
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        threads.clear()
        threads_url.clear()

    end = datetime.datetime.now()
    print("豆瓣爬虫爬取结束...")
    print('运行时间：{time}'.format(time=(end - start)))
