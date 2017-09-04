#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import re
import logging
import openpyxl
import threading
import queue
import datetime
import os


def getHTMLText(url):
    """
    下载目标网页源码
    """
    try:
        r = requests.get(url)
        r.raise_for_status()
        r.encoding = 'utf-8'
        return r.text
    except Exception:
        logger.exception("Download HTML Text failed")


def getHTMKUrl(url):
    """
    提取萧山、滨江待爬取的租房网页链接
    """
    html = getHTMLText(url)
    soup = BeautifulSoup(html, 'html.parser')
    try:
        pagination = soup.find(attrs={'class': 'pagination_group_a'})
        page = pagination.find_all('a')
        for i in range(len(page)):
            page_url = page[i].attrs['href']
            url_queue.put(page_url)
    except Exception:
        logger.exception("Site link extraction failed")


def getContentExtraction(url_path, num):
    """
    提取每个网页中中的租房链接、租房标题、租金、小区名、户型、平方、朝向、更新时间、装修的信息并保存到Excel
    """
    aims_url = 'https://hz.lianjia.com' + url_path
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/lianjia_renting.xlsx'
    renting_list = []
    html = getHTMLText(aims_url)
    soup = BeautifulSoup(html, 'html.parser')
    try:
        ul = soup.find(attrs={'class': 'house-lst'})
        renting_info = ul.find_all(attrs={'class': 'info-panel'})
        price = ul.find_all(attrs={'class': 'price'})
        price_pre = ul.find_all(attrs={'class': 'price-pre'})
        where = ul.find_all(attrs={'class': 'where'})
        view_label = ul.find_all(attrs={'class': 'view-label left'})
        # print(where[0].span.text.split()[0])
        for i in range(len(renting_info)):
            renting_data = {'租房链接': '', '租房标题': '', '租金': '', '小区名': '', '户型': '', '平方': '', '朝向': '', '更新时间': '', '装修': '', }
            # 租房链接、租房标题
            renting_url = renting_info[i].a.attrs['title']
            renting_title = renting_info[i].a.attrs['href']
            # 租金
            rent = price[i].span.text
            # 小区名、户型、平方、朝向
            community_name = where[i].span.text.split()[0]
            huxing = re.findall(r'\d室\d厅', str(where[i]))[0]
            square = re.findall(r'\d*平米', str(where[i]))[0]
            direction = re.findall(r'[东南西北]', str(where[i]))
            orientation = ''
            for dire in direction:
                orientation += dire
            # 装修程度
            decoration_list = re.findall(r'精装修', str(view_label[i]))
            if decoration_list:
                decoration = decoration_list[0]
            else:
                decoration = ''
            # 更新时间
            update_time = price_pre[i].text

            renting_data['更新时间'] = update_time
            renting_data['租房链接'] = renting_url
            renting_data['租房标题'] = renting_title
            renting_data['租金'] = rent
            renting_data['小区名'] = community_name
            renting_data['户型'] = huxing
            renting_data['平方'] = square
            renting_data['朝向'] = decoration

            renting_list.append(renting_data)

        mutex.acquire()
        if not os.path.exists(pata):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for x in range(len(renting_list)):
                sheet["A%d" % (x + 1)].value = renting_list[x]['更新时间']
                sheet["B%d" % (x + 1)].value = renting_list[x]['租房标题']
                sheet["C%d" % (x + 1)].value = renting_list[x]['租金']
                sheet["D%d" % (x + 1)].value = renting_list[x]['小区名']
                sheet["E%d" % (x + 1)].value = renting_list[x]['户型']
                sheet["F%d" % (x + 1)].value = renting_list[x]['平方']
                sheet["G%d" % (x + 1)].value = renting_list[x]['朝向']
                sheet["H%d" % (x + 1)].value = renting_list[x]['租房链接']
            workbook.save(pata)
        else:
            workbook = openpyxl.load_workbook(pata)
            sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
            row = sheet.max_row
            for x in range(len(renting_list)):
                sheet["A%d" % (row + x + 1)].value = renting_list[x]['更新时间']
                sheet["B%d" % (row + x + 1)].value = renting_list[x]['租房标题']
                sheet["C%d" % (row + x + 1)].value = renting_list[x]['租金']
                sheet["D%d" % (row + x + 1)].value = renting_list[x]['小区名']
                sheet["E%d" % (row + x + 1)].value = renting_list[x]['户型']
                sheet["F%d" % (row + x + 1)].value = renting_list[x]['平方']
                sheet["G%d" % (row + x + 1)].value = renting_list[x]['朝向']
                sheet["H%d" % (row + x + 1)].value = renting_list[x]['租房链接']
            workbook.save(pata)
        print("\r第{number}页数据爬取结束".format(number=num, end=""))
        mutex.release()

    except Exception:
        logger.exception("Extract page {number} message failed".format(number=num))

if __name__ == '__main__':
    logging.basicConfig(level=logging.WARNING,
                        format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='/Users/wangjiacan/Desktop/代码/log/lianjia_renting.txt',
                        filemode='a')

    logger = logging.getLogger()
    url_queue = queue.Queue()

    lianjia_url = ['https://hz.lianjia.com/zufang/binjiang/', 'https://hz.lianjia.com/zufang/xiaoshan/']
    number = 0
    threads = []
    mutex = threading.Lock()

    print("链家租房爬虫爬取开始...")
    start = datetime.datetime.now()
    for lj_url in lianjia_url:
        getHTMKUrl(lj_url)
    for t in range(url_queue.qsize()):
        number += 1
        threads.append(threading.Thread(target=getContentExtraction, args=(url_queue.get(), number)))
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    end = datetime.datetime.now()
    print("\n链家租房爬虫爬取结束...")
    print('运行时间：{time}'.format(time=(end - start)))
