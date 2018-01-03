#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import openpyxl
import os


def dataProcessing(fileLocation, title, tableName):
    """
    转化写入数据函数
    输入字段说明：
    title：待转化文本列表，列表类型。
    输出字段说明：
    writeContentList：excel模块可写入的文本列表，列表类型
    """
    formFields = []
    writeContentList = []
    contentList = []
    positionList = []
    newContentList = []
    completeContentList = []

    columnList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'O']
    columnNum = len(title[0])
    rowsNum = len(title)
    newColumnList = columnList[0:columnNum]

    # 获取表格表头字段
    readWorkbook = openpyxl.load_workbook(fileLocation)
    tableNameList = readWorkbook.get_sheet_names()
    position = tableNameList.index(tableName)
    readSheet = readWorkbook.get_sheet_by_name(readWorkbook.get_sheet_names()[position])
    column = readSheet.max_column
    row = readSheet.max_row

    for num in range(column):
        num += 1
        formFields.append(readSheet.cell(row=1,column=num).value)


    # 处理数据：按照数据的数量与个数，生成column与rows数据
    for num_1 in range(rowsNum):
        newRow = num_1 + 1
        for num_2 in range(columnNum):
            writeContent = {'column': '', 'rows': '', 'content': '', }
            column = newColumnList[num_2]
            writeContent['column'] = column
            writeContent['rows'] = newRow + row
            writeContentList.append(writeContent)

    # 获取输入数据中字段的排序顺序
    for key in title[0]:
        position = formFields.index(key)
        positionList.append(position)


    # 获取输入数据中的内容
    for num in range(len(title)):
        contentInfo = []
        for key in title[num]:
            content = title[num][key]
            contentInfo.append(content)
        contentList.append(contentInfo)


    # 处理数据：按照表头的顺序排序输入数据内容
    for contentInfo in contentList:
        newContentInfo = []
        for num in range(len(contentInfo)):
            content = contentInfo[num]
            position = positionList[num]

            if position == num:
                newContentInfo.append(content)
            elif position > num:
                newContentInfo.append(content)
            else:
                newContentInfo.insert(position, content)

        newContentList.append(newContentInfo)


    # 处理数据：根据排序好的数据整理成可编辑的表格
    for contentList in newContentList:
        for content in contentList:
            completeContentList.append(content)


    for num in range(len(writeContentList)):
        writeContentList[num]['content'] = completeContentList[num]

    return writeContentList


def writeExcel(fileLocation, title, tableName):
    """
    excel写入函数
    输入字段说明：
    fileLocation：excel存放路径，字符串类型
    tableName：excel子表表名，字符串类型
    title：excel模块需写入的文本列表，列表类型。
    """
    writeContentList = dataProcessing(fileLocation, title, tableName)

    if not os.path.exists(fileLocation):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = tableName
        workbook.save(fileLocation)

    existedWorkbook = openpyxl.load_workbook(fileLocation)
    tableNameList = existedWorkbook.get_sheet_names()
    if tableName in tableNameList:
        position = tableNameList.index(tableName)
        existedSheet = existedWorkbook.get_sheet_by_name(existedWorkbook.get_sheet_names()[position])
        row = existedSheet.max_row

        if row == 1:
            row = 0

        for i in range(len(writeContentList)):
            column = writeContentList[i]['column']
            content = writeContentList[i]['content']
            rows = writeContentList[i]['rows'] + row
            existedSheet["{column}{rows}".format(column=column, rows=rows)].value = content
    else:
        newSheet = existedWorkbook.create_sheet()
        newSheet.title = tableName
        for i in range(len(writeContentList)):
            column = writeContentList[i]['column']
            content = writeContentList[i]['content']
            rows = writeContentList[i]['rows']
            newSheet["{column}{rows}".format(column=column, rows=rows)].value = content

    existedWorkbook.save(fileLocation)


def readExcel(fileLocation, tableName, columnCount, excelType):
    """
    excel读取函数
    输入字段说明：
    fileLocation：excel存放路径，字符串类型
    tableName：excel子表表名，字符串类型
    columnCount：读取的列数列表，列表类型
    excelType：读取的excel标题栏，列表类型
    输出字段说明：
    excelInfoList：返回已读取的数据列表，列表类型
    """
    excelInfoList = []

    readWorkbook = openpyxl.load_workbook(fileLocation)
    tableNameList = readWorkbook.get_sheet_names()
    position = tableNameList.index(tableName)
    readSheet = readWorkbook.get_sheet_by_name(readWorkbook.get_sheet_names()[position])

    for row in readSheet.rows:
        rowsInfoList = []

        for column in columnCount:
            rowsInfoList.append(row[column].value)
        excelInfoData = dict(zip(excelType, rowsInfoList))

        excelInfoList.append(excelInfoData)

    readWorkbook.save(fileLocation)

    return excelInfoList

def coverExcel(fileLocation, tableName, title):
    """
    excel覆盖函数
    输入字段说明：
    fileLocation：excel存放路径，字符串类型
    tableName：excel子表表名，字符串类型
    title：excel模块需写入的文本列表，列表类型。
    """
    writeContentList = dataProcessing(fileLocation, title, tableName)

    coverWorkbook = openpyxl.load_workbook(fileLocation)
    tableNameList = coverWorkbook.get_sheet_names()
    position = tableNameList.index(tableName)
    coverSheet = coverWorkbook.get_sheet_by_name(coverWorkbook.get_sheet_names()[position])

    coverWorkbook.remove_sheet(coverSheet)

    newSheet = coverWorkbook.create_sheet(tableName)

    for i in range(len(writeContentList)):
        column = writeContentList[i]['column']
        content = writeContentList[i]['content']
        rows = writeContentList[i]['rows']
        newSheet["{column}{rows}".format(column=column, rows=rows)].value = content

    coverWorkbook.save(fileLocation)
