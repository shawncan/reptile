#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import openpyxl
import os
import sys
import configparser
import pymysql
import re


class Settlement(object):
    #===========================================================================
    # '''
    # '''
    #===========================================================================
    def __init__(self):
        conf = configparser.ConfigParser()
        conf.read("/Users/wangjiacan/Desktop/sourceCode/configurationFile/localConfiguration.ini")

        self.host = conf.get("localServer", "host")
        self.port = int(conf.get("localServer", "port"))
        self.user = conf.get("localServer", "user")
        self.password = conf.get("localServer", "password")
        self.dbName = conf.get("localServer", "dbname")

        self.fileAddress = ''
        self.outputName = ''
        self.tableName = ''
        self.excelInfoList = []


    def getCursor(self):
        # 建立数据库链接
        self.db = pymysql.connect(host=self.host, port=self.port, user=self.user, password=self.password, db=self.dbName, charset='utf8')

        # 创建游标对象
        cursor = self.db.cursor()

        # 返回游标对象
        return cursor


    def queryOperation(self,sql):
        # 建立连接获取游标对象
        cur = self.getCursor()

        # 执行SQL语句
        cur.execute(sql)

        # 获取查询数据
        # fetch*
        # all 所有数据, one 取结果的一行, many(size),去size行
        dataList = cur.fetchall()

        # 关闭游标
        cur.close()

        # 关闭数据库连接
        self.db.close()

        # 返回查询数据
        return dataList

    def readExcel(self):
        """
        """
        readWorkbook = openpyxl.load_workbook(self.fileAddress)

        self.tableName = input("输入需要导入的表名：")
        readSheet = readWorkbook[self.tableName]

        order = 0
        for row in readSheet.rows:
            order +=1

            if order == 1:
                continue

            excelInfo = {'标注人': '', '团队长': '', '账号': '', '正确率': '', '数量': '', '价格': '',}

            user = row[0].value
            head = row[1].value
            details = row[2].value
            correctRate = row[3].value

            if not user:
                continue

            excelInfo['标注人'] = user
            excelInfo['团队长'] = head
            excelInfo['账号'] = re.findall(r"[A-Za-z0-9]{13}", details)[0]
            excelInfo['正确率'] = correctRate

            self.excelInfoList.append(excelInfo)


    def readQuantity(self):
        # 查询sql语句
        inquire_sql = """
        select quantity from azy_taskData WHERE account = "{account}" 
        """

        for infoData in self.excelInfoList:
            account = infoData["账号"]

            searchResult = self.queryOperation(inquire_sql.format(account=account))

            if not searchResult:
                quantity = 0
            else:
                quantity = searchResult[0][0]

            infoData["数量"] = quantity


    def calculateCost(self):
        billingType = int(input("输入结算类型："))

        for infoData in self.excelInfoList:
            calculateCost = 0

            quantity = infoData["数量"]
            correctRate = infoData["正确率"]

            if billingType == 1:
                if correctRate >= 96:
                    calculateCost = float(quantity) * 0.02
                elif correctRate >=93:
                    calculateCost = float(quantity) * 0.015
            elif billingType == 2:
                calculateCost = float(quantity) * 0.015
            else:
                print("输入结算类型错误，请重新运行。")
                sys.exit()

            infoData["价格"] = calculateCost


    def exportExcel(self):
        if not os.path.exists(self.outputName):
            workbook = openpyxl.Workbook()
            workbook.save(self.outputName)

        exportWorkbook = openpyxl.load_workbook(self.outputName)

        newSheet = exportWorkbook.active

        for i in range(len(self.excelInfoList)):

            user = self.excelInfoList[i]['标注人']
            head = self.excelInfoList[i]['团队长']
            details = self.excelInfoList[i]['账号']
            correctRate = self.excelInfoList[i]['正确率']
            quantity = self.excelInfoList[i]['数量']
            calculateCost = self.excelInfoList[i]['价格']

            exportInfoList = [user, head, details, correctRate, quantity, calculateCost]

            newSheet.append(exportInfoList)


        exportWorkbook.save(self.outputName)


    def start(self):
        self.fileAddress = input("输入需要结算的文件名：")

        if not os.path.exists(self.fileAddress):
            print("{name}文件未找到，请确认：1.文件名是否输入正确。2.导入文件与脚本文件是否放在一起。".format(name=self.fileAddress))
            sys.exit()

        self.readExcel()
        self.readQuantity()
        self.calculateCost()

        enable = True

        while enable:
            self.outputName = input("输入需要输出的文件名：")

            if os.path.exists(self.outputName):
                print("{name}已存在，请确认：1.文件名是否输入正确。".format(name=self.outputName))
                continue

            enable = False

        self.exportExcel()


if __name__ == '__main__':
    extract = Settlement()
    extract.start()