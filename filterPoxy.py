#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
import openpyxl


def verification(ip):
    """
    检测代理的ip是否可用
    """
    status = True
    url = 'https://httpbin.org/get?show_env=1'
    headers = {'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Mobile Safari/537.36'}

    try:
        proxies = {
            "http": "http://" + ip,
            "https": "http://" + ip
            }
        r = requests.get(url, headers=headers, proxies=proxies, timeout=20)
        r.raise_for_status()

    except Exception:
        status = False

    return status


def ipExtract():
    success_count = 0
    failure_count = 0
    available = [{'ip': 'ip', '类型': '类型', '验证时间': '验证时间', }]
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/agentPool.xlsx'
    page_workbook = openpyxl.load_workbook(pata)
    page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0])
    row = page_sheet.max_row

    for i in range(row):
        proxy_data = {'ip': '', '类型': '', '验证时间': '', }

        if i == 0:
            continue
        ip = page_sheet["A%d" % (i + 1)].value
        type = page_sheet["B%d" % (i + 1)].value
        verification_time = page_sheet["C%d" % (i + 1)].value

        link_status = verification(ip)
        if link_status:
            success_count += 1
            proxy_data['ip'] = ip
            proxy_data['类型'] = type
            proxy_data['验证时间'] = verification_time

            available.append(proxy_data)
        else:
            failure_count += 1

    # 删除表
    page_workbook.remove_sheet(page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0]))

    # 新建表
    new_sheet = page_workbook.create_sheet()

    for i in range(len(available)):
        new_sheet["A%d" % (i + 1)].value = available[i]['ip']
        new_sheet["B%d" % (i + 1)].value = available[i]['类型']
        new_sheet["C%d" % (i + 1)].value = available[i]['验证时间']

    page_workbook.save(pata)
    print("ip extracted successfully：{}valid data，{}failure data".format(success_count, failure_count))


if __name__ == '__main__':
    ipExtract()
