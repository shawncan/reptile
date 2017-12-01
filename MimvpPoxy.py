#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
import openpyxl
import os

def getProxy():
    """
    下载代理的ip
    """
    api_url = '请填写米扑代理提取链接'
    result = requests.get(api_url)
    proxy_json = result.json()
    return proxy_json['result']


def verification(ip):
    """
    检测代理的ip是否可用
    """
    verificationStatus = 0
    url = 'https://httpbin.org/get?show_env=1'
    headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:57.0) Gecko/20100101 Firefox/57.0'}
    proxies = {
        "http": "http://" + ip,
        "https": "http://" + ip
    }
    try:
        r = requests.get(url, headers=headers, proxies=proxies, timeout=20)
        r.raise_for_status()
    except:
        verificationStatus = 1

    return verificationStatus



def writeExcel(poxy_list):
    """
    把提取出来的信息写入excel
    """
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/agentPool.xlsx'
    title = ['ip', '类型', '验证时间',]

    if not os.path.exists(pata):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"].value = title[0]
        sheet["B1"].value = title[1]
        sheet["C1"].value = title[2]
        workbook.save(pata)

    page_workbook = openpyxl.load_workbook(pata)
    page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0])
    row = page_sheet.max_row
    for i in range(len(poxy_list)):
        page_sheet["A%d" % (row + i + 1)].value = poxy_list[i]['ip:port']
        page_sheet["B%d" % (row + i + 1)].value = poxy_list[i]['http_type']
        page_sheet["C%d" % (row + i + 1)].value = poxy_list[i]['check_dtime']
    page_workbook.save(pata)


def getContentExtraction():
    success_count = 0
    failure_count = 0
    poxyList = []
    proxyInfo = getProxy()
    for i in range(len(proxyInfo)):
        ip = proxyInfo[i]['ip:port']
        status = verification(ip)

        if status == 0:
            success_count += 1
            print("{ip}筛选成功".format(ip=ip))
            poxyList.append(proxyInfo[i])
        else:
            failure_count += 1
            print("{ip}筛选失败".format(ip=ip))

    writeExcel(poxyList)
    print("代理ip成功：爬取成功{}条数据，爬取失败{}条数据".format(success_count, failure_count))







if __name__ == '__main__':
    # url_queue = queue.Queue()
    # print("代理ip提取开始...")
    # getProxy()
    # print("代理ip下载完成，开始检测ip可用性")
    # verification()
    # print("代理ip提取结束...")
    getContentExtraction()
