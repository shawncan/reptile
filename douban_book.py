#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import re
import logging
import urllib.parse
import openpyxl
import threading
import queue
import datetime
import time
import os


def getHTMLText(url):
    try:
        r = requests.get(url)
        r.raise_for_status()
        r.encoding = 'utf-8'
        return r.text
    except Exception:
        logger.exception("Download HTML Text failed")


def getHTMLUrl(url):
    html = getHTMLText(url)
    soup = BeautifulSoup(html, 'html.parser')
    try:
        paginator = soup.find(attrs={'class': 'paginator'})
        page = paginator.find_all('a')

        for i in range(len(page)):
            next = page[i].text
            page_url = page[i].attrs['href']
            if next == "后页>":
                continue
            url_queue.put(page_url)
    except Exception:
        logger.exception("Site link extraction failed")


def getContentExtraction(url):
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/duoban_book.xlsx'
    book_list = []
    html = getHTMLText(url)
    soup = BeautifulSoup(html, 'html.parser')

    try:
        indent = soup.find(attrs={'class': 'indent'})
        pl2 = indent.find_all(attrs={'class': 'pl2'})
        pl = indent.find_all('p', attrs={'class': 'pl'})
        rating_nums = indent.find_all(attrs={'class': 'rating_nums'})
        span_pl = indent.find_all('span', attrs={'class': 'pl'})
        item = indent.find_all('tr', attrs={'class': 'item'})

        for i in range(len(pl2)):
            book_data = {'书名': '', '作者': '', '评分': '', '评价人数': '', '评语': '', }

            book_name = pl2[i].a.attrs['title']
            book_information = pl[i].text
            author = book_information.split('/')[0]
            score = rating_nums[i].text
            number_of_comments = re.findall(r'\d*人评价', str(span_pl[i]))[0]

            inq = re.findall(r'\<span class\=\"inq\"\>.*\<\/span\>', str(item[i]))
            if inq:
                comments = inq[0][18:-7]
            else:
                comments = ""

            book_data['书名'] = book_name
            book_data['作者'] = author
            book_data['评分'] = score
            book_data['评价人数'] = number_of_comments
            book_data['评语'] = comments

            book_list.append(book_data)

        mutex.acquire()
        if not os.path.exists(pata):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for i in range(len(book_list)):
                sheet["A%d" % (i + 1)].value = book_list[i]['书名']
                sheet["B%d" % (i + 1)].value = book_list[i]['作者']
                sheet["C%d" % (i + 1)].value = book_list[i]['评分']
                sheet["D%d" % (i + 1)].value = book_list[i]['评价人数']
                sheet["E%d" % (i + 1)].value = book_list[i]['评语']
            workbook.save(pata)
        else:
            workbook = openpyxl.load_workbook(pata)
            sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
            row = sheet.max_row
            for i in range(len(book_list)):
                sheet["A%d" % (row + i + 1)].value = book_list[i]['书名']
                sheet["B%d" % (row + i + 1)].value = book_list[i]['作者']
                sheet["C%d" % (row + i + 1)].value = book_list[i]['评分']
                sheet["D%d" % (row + i + 1)].value = book_list[i]['评价人数']
                sheet["E%d" % (row + i + 1)].value = book_list[i]['评语']
            workbook.save(pata)
        mutex.release()
    except Exception:
        logger.exception("Extract page {number} message failed".format(number=number))


if __name__ == '__main__':
    logging.basicConfig(level=logging.WARNING,
                        format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='/Users/wangjiacan/Desktop/代码/log/DBTS.txt',
                        filemode='a')

    logger = logging.getLogger()
    url_queue = queue.Queue()

    number = 0

    print("豆瓣爬虫爬取开始...")
    start = datetime.datetime.now()
    initial_url = 'https://book.douban.com/top250?start=0'
    url_queue.put(initial_url)
    getHTMLUrl(initial_url)
    mutex = threading.Lock()

    for i in range(url_queue.qsize()):
        number += 1
        t = threading.Thread(target=getContentExtraction, args=(url_queue.get(),))
        t.start()
        print("\r第{number}页数据开始爬取".format(number=number, end=""))

    end = datetime.datetime.now()
    time.sleep(5)
    print("\n豆瓣爬虫爬取结束...")
    print('运行时间：{time}'.format(time=(end - start)))
