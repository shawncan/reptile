#!/usr/local/Cellar/python3
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import re
import logging
import openpyxl
import os
import json


def getHTMLText(url, code = 'utf-8'):
    """
    下载目标网页源码
    """
    proxyInfo = []
    html = ''
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:57.0) Gecko/20100101 Firefox/57.0',
    }

    readExcel(proxyInfo)

    for num in range(len(proxyInfo)):
        if num == 0:
            continue

        proxyIP = proxyInfo[num]['ip']
        proxies = {
            "http": "http://" + proxyIP,
            "https": "http://" + proxyIP,
        }

        try:
            r = requests.get(url, headers=headers, proxies=proxies, timeout=20)
            r.raise_for_status()
            r.encoding = code
            html = r.text
        except Exception:
            html = None

        if html == None:
            proxyInfo.remove(proxyInfo[num])
            continue
        else:
            break

    writrProxyExcel(proxyInfo)
    return html


def readExcel(proxyInfo_list):
    """
    读取表格中代理IP
    """
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/agentPool.xlsx'
    page_workbook = openpyxl.load_workbook(pata)
    page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0])
    for row in page_sheet.rows:
        proxyData = {'ip': '', '类型': '', '验证时间': '', }
        ip = row[0].value
        httpType = row[1].value
        checkDtime = row[2].value

        proxyData["ip"] = ip
        proxyData["类型"] = httpType
        proxyData["验证时间"] = checkDtime

        proxyInfo_list.append(proxyData)
    page_workbook.save(pata)


def writrProxyExcel(proxyInfo_list):
    """
    写入使用后有效代理IP
    """
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/agentPool.xlsx'
    page_workbook = openpyxl.load_workbook(pata)

    # 删除表
    page_workbook.remove_sheet(page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[0]))

    # 新建表
    new_sheet = page_workbook.create_sheet()

    for i in range(len(proxyInfo_list)):
        new_sheet["A%d" % (i + 1)].value = proxyInfo_list[i]['ip']
        new_sheet["B%d" % (i + 1)].value = proxyInfo_list[i]['类型']
        new_sheet["C%d" % (i + 1)].value = proxyInfo_list[i]['验证时间']

    page_workbook.save(pata)


def getShopId(productList, shop_list):
    """
    获取店铺sku商品id
    """
    soup = BeautifulSoup(productList, 'html.parser')
    glWarpClearfix = soup.find(attrs={'class': 'gl-warp clearfix'})
    glItem = glWarpClearfix.find_all(attrs={'class': 'gl-item'})
    for shop in glItem:
        shopId = shop.attrs['data-sku']
        shop_list.append(shopId)


def writeExcel(comment_list, shop_id):
    """
    把提取出来的信息写入excel
    """
    pata = '/Users/wangjiacan/Desktop/shawn/爬取资料/jdShopComment.xlsx'
    title = {'商品类型': '商品类型', '用户名': '用户名', '会员等级': '会员等级', '评价时间': '评价时间', '评价内容': '评价内容',
             '追评时间': '追评时间', '追评内容': '追评内容', '评价客户端': '评价客户端','评价类型': '评价类型',}

    if not os.path.exists(pata):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = shop_id
        sheet["A1"].value = title['商品类型']
        sheet["B1"].value = title['用户名']
        sheet["C1"].value = title['会员等级']
        sheet["D1"].value = title['评价时间']
        sheet["E1"].value = title['评价内容']
        sheet["F1"].value = title['追评时间']
        sheet["G1"].value = title['追评内容']
        sheet["H1"].value = title['评价客户端']
        sheet["I1"].value = title['评价类型']
        workbook.save(pata)

    page_workbook = openpyxl.load_workbook(pata)
    table_Name = page_workbook.get_sheet_names()
    if shop_id in table_Name:
        position = table_Name.index(shop_id)
        page_sheet = page_workbook.get_sheet_by_name(page_workbook.get_sheet_names()[position])
        row = page_sheet.max_row
        for i in range(len(comment_list)):
            page_sheet["A%d" % (row + i + 1)].value = comment_list[i]['商品类型']
            page_sheet["B%d" % (row + i + 1)].value = comment_list[i]['用户名']
            page_sheet["C%d" % (row + i + 1)].value = comment_list[i]['会员等级']
            page_sheet["D%d" % (row + i + 1)].value = comment_list[i]['评价时间']
            page_sheet["E%d" % (row + i + 1)].value = comment_list[i]['评价内容']
            page_sheet["F%d" % (row + i + 1)].value = comment_list[i]['追评时间']
            page_sheet["G%d" % (row + i + 1)].value = comment_list[i]['追评内容']
            page_sheet["H%d" % (row + i + 1)].value = comment_list[i]['评价客户端']
            page_sheet["I%d" % (row + i + 1)].value = comment_list[i]['评价类型']
    else:
        comment_list.insert(0, title)
        new_sheet = page_workbook.create_sheet()
        new_sheet.title = shop_id
        for i in range(len(comment_list)):
            new_sheet["A%d" % (i + 1)].value = comment_list[i]['商品类型']
            new_sheet["B%d" % (i + 1)].value = comment_list[i]['用户名']
            new_sheet["C%d" % (i + 1)].value = comment_list[i]['会员等级']
            new_sheet["D%d" % (i + 1)].value = comment_list[i]['评价时间']
            new_sheet["E%d" % (i + 1)].value = comment_list[i]['评价内容']
            new_sheet["F%d" % (i + 1)].value = comment_list[i]['追评时间']
            new_sheet["G%d" % (i + 1)].value = comment_list[i]['追评内容']
            new_sheet["H%d" % (i + 1)].value = comment_list[i]['评价客户端']
            new_sheet["I%d" % (i + 1)].value = comment_list[i]['评价类型']
    page_workbook.save(pata)


def commentProcessing(contentInfo):
    """
    处理评论输出中文评论
    """
    symbol_list = re.findall(r"&[a-z]*;", contentInfo)
    if symbol_list:
        for symbol in set(symbol_list):
            if symbol == "&hellip;":
                new_symbol = "..."
            elif symbol == "&ldquo;":
                new_symbol = "'"
            elif symbol == "&rdquo;":
                new_symbol = "'"
            else:
                new_symbol = " "
            new_content = re.sub(symbol, new_symbol, contentInfo)
            contentInfo = new_content
    return contentInfo


def getContentExtraction(buyersomments, score, shop_id):
    """
    提取评论链接中的信息
    """
    jdkkComment_list = []

    if int(score) == 1:
        commentType = '差评'
    else:
        commentType = '中评'

    for num in range(len(buyersomments)):
        kkComment_data = {'商品类型': '', '用户名': '', '会员等级': '', '评价时间': '', '评价内容': '','追评时间': '',
                          '追评内容': '', '评价客户端': '', '评价类型': '',}

        # 商品类型
        productColor = buyersomments[num]["productColor"]
        # 用户名
        nickname = buyersomments[num]["nickname"]
        # 会员等级
        userLevelName = buyersomments[num]["userLevelName"]

        # 第一次评价时间
        creationTime = buyersomments[num]["creationTime"]
        # 第一次评价内容
        content = buyersomments[num]["content"]
        modify_content = commentProcessing(content)

        # 第二次评价信息
        created = ''
        modifyMostComment = ''
        if "afterUserComment" in buyersomments[num].keys():
            afterUserComment = buyersomments[num]["afterUserComment"]
            # 第二次评价时间
            created = afterUserComment["created"]
            # 第二次评价内容
            hAfterUserComment = afterUserComment["hAfterUserComment"]
            mostComment = hAfterUserComment["content"]
            modifyMostComment = commentProcessing(mostComment)

        # 评价客户端
        userClientShow = buyersomments[num]['userClientShow']

        kkComment_data['商品类型'] = productColor
        kkComment_data['用户名'] = nickname
        kkComment_data['会员等级'] = userLevelName
        kkComment_data['评价时间'] = creationTime
        kkComment_data['评价内容'] = modify_content
        kkComment_data['追评时间'] = created
        kkComment_data['追评内容'] = modifyMostComment
        kkComment_data['评价客户端'] = userClientShow
        kkComment_data['评价类型'] = commentType

        jdkkComment_list.append(kkComment_data)

    writeExcel(jdkkComment_list, shop_id)


def getEvaluationUrl(shop_id):
    """
    下载并写入商品id下的中评、差评
    """
    # score：2为中评，1为差评
    scoreList = ['1', '2']
    for score in scoreList:
        enable = True
        page = 0

        if int(score) == 1:
            commentType = '差评'
        else:
            commentType = '中评'

        while enable:
            evaluationUrl = 'https://sclub.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98vv445&' \
                            'productId={productId}&score={score}&sortType=5&page={page}&pageSize=10&isShadowSku=0&rid=0&fold=1'.format \
                (productId=shop_id, score=score, page=page)
            html = getHTMLText(evaluationUrl, 'GBK')

            htmlInfo = html[25:-2]
            returnJosn = json.loads(htmlInfo)
            buyersCommentsInfo = returnJosn['comments']
            if buyersCommentsInfo:
                getContentExtraction(buyersCommentsInfo, score, shop_id)
                print("{commentType}第{page}页评论爬取完成".format(commentType=commentType, page=page))
            else:
                enable = False
            page += 1

def Start():
    """
    搜索输入的店铺名称并获取该店铺下所有sku商品id
    """
    shopList = []
    shopName = '控客京东自营旗舰店'

    # 获取店铺下所有sku的商品id
    shopSearchUrl = 'https://search.jd.com/Search?keyword={shopName}&enc=utf-8&pvid=75680dd04a8c4c1692304f1fee4e591a'.format\
        (shopName=shopName, )

    shopSearchHtml = getHTMLText(shopSearchUrl)
    getShopId(shopSearchHtml, shopList)

    # 获取单sku商品下的所有评论
    for shopid in shopList:
        print("{shopName}的商品：{shopid}".format(shopName=shopName, shopid=shopid))
        getEvaluationUrl(shopid)
    print("评论完全爬取完成")


if __name__ == '__main__':
    logging.basicConfig(level=logging.WARNING,
                        format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='/Users/wangjiacan/Desktop/代码/log/jdShopComment.txt',
                        filemode='a')

    logger = logging.getLogger()

    Start()